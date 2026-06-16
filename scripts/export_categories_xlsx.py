"""Выгрузка Stop-Piramida по категориям в Excel — для раздачи студентам на разметку.

На выходе (в data/processed/):
  categories_xlsx/<категория>.xlsx   — по файлу на каждую из 17 категорий
  stop_piramida_by_category.xlsx     — единая книга, лист на категорию

Колонки: метаданные + транскрипт + авто-разметка (fraud_type/label/risk от системы)
+ ПУСТЫЕ колонки для ручной разметки студентом (student_label/fraud_type/evidence/notes).

Запуск:
    python -m scripts.export_categories_xlsx
"""

from __future__ import annotations

import collections
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = "data/processed/stop_piramida.jsonl"
OUT_DIR = "data/processed/categories_xlsx"
COMBINED = "data/processed/stop_piramida_by_category.xlsx"

# (заголовок, ключ-извлечение, ширина, перенос)
COLUMNS = [
    ("id", lambda r: r.get("id"), 16, False),
    ("category", lambda r: r.get("case_type"), 20, False),
    ("fraud_type (авто)", lambda r: r.get("fraud_type"), 22, False),
    ("label (авто)", lambda r: r.get("label"), 12, False),
    ("risk_level", lambda r: r.get("risk_level"), 11, False),
    ("risk_score", lambda r: r.get("risk_score"), 10, False),
    ("title", lambda r: r.get("title"), 28, True),
    ("description", lambda r: r.get("description"), 40, True),
    ("transcript", lambda r: r.get("transcript"), 80, True),
    ("risk_signals", lambda r: ", ".join(r.get("risk_signals") or []), 30, True),
    ("domains", lambda r: ", ".join((r.get("entities") or {}).get("domains") or []), 18, True),
    ("telegram", lambda r: ", ".join((r.get("entities") or {}).get("telegram_usernames") or []), 16, True),
    ("promo_codes", lambda r: ", ".join((r.get("entities") or {}).get("promo_codes") or []), 14, True),
    ("organizations", lambda r: ", ".join((r.get("entities") or {}).get("organizations") or []), 22, True),
    ("page_url", lambda r: r.get("url"), 30, False),
    ("vimeo / media", lambda r: r.get("media_path") or "", 30, False),
    ("review_status", lambda r: r.get("review_status"), 18, False),
    # пустые колонки для студента
    ("✍ student_label", lambda r: "", 16, False),
    ("✍ student_fraud_type", lambda r: "", 20, False),
    ("✍ student_evidence", lambda r: "", 30, True),
    ("✍ student_notes", lambda r: "", 30, True),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
STUDENT_FILL = PatternFill("solid", fgColor="3B5021")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def fill_sheet(ws, rows: list[dict]) -> None:
    # заголовок
    for col, (title, _, width, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = HEADER_FONT
        c.fill = STUDENT_FILL if title.startswith("✍") else HEADER_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    # данные
    for ri, rec in enumerate(rows, 2):
        for col, (_, getter, _, wrap) in enumerate(COLUMNS, 1):
            c = ws.cell(row=ri, column=col, value=getter(rec))
            if wrap:
                c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"


def safe_sheet_name(name: str) -> str:
    for ch in r'[]:*?/\\':
        name = name.replace(ch, "_")
    return name[:31]


def main() -> None:
    records = [json.loads(line) for line in open(SRC, encoding="utf-8") if line.strip()]
    by_cat: dict[str, list[dict]] = collections.defaultdict(list)
    for r in records:
        by_cat[r.get("case_type") or "unknown"].append(r)

    os.makedirs(OUT_DIR, exist_ok=True)

    # отдельный файл на категорию
    for cat, rows in sorted(by_cat.items()):
        wb = Workbook()
        fill_sheet(wb.active, rows)
        wb.active.title = safe_sheet_name(cat)
        wb.save(os.path.join(OUT_DIR, f"{cat}.xlsx"))

    # единая книга — лист на категорию
    wb = Workbook()
    wb.remove(wb.active)
    for cat, rows in sorted(by_cat.items()):
        ws = wb.create_sheet(safe_sheet_name(cat))
        fill_sheet(ws, rows)
    wb.save(COMBINED)

    print(f"Категорий: {len(by_cat)}")
    for cat, rows in sorted(by_cat.items()):
        tr = sum(1 for r in rows if r.get("transcript"))
        print(f"  {cat:32} {len(rows):3} видео ({tr} с транскриптом)")
    print(f"\nОтдельные файлы → {OUT_DIR}/<категория>.xlsx")
    print(f"Единая книга    → {COMBINED}")


if __name__ == "__main__":
    main()
